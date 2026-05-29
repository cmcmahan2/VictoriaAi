"""
Spreadsheet export for discovered leads.

Writes both .xlsx (with colour-coded rows) and .csv to the output directory.
Returns (xlsx_path, csv_path) as strings.
"""

import csv
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

COLUMNS = [
    ("rank",          "Rank"),
    ("name",          "Name"),
    ("phone",         "Phone"),
    ("address",       "Address"),
    ("city",          "City"),
    ("neighborhood",  "Neighborhood"),
    ("category",      "Category"),
    ("existing_website", "Website"),
    ("score",         "Score"),
    ("weakness_flags", "Weakness Flags"),
    ("google_maps_url", "Google Maps URL"),
    ("rating",        "Rating"),
    ("review_count",  "Reviews"),
    ("source",        "Source"),
]

_FILL_RED    = "FFE6E6"
_FILL_YELLOW = "FFFDE6"
_FILL_GREEN  = "E6FFE9"
_FILL_HEADER = "1E2D3D"


def _cell_value(lead: dict, key: str):
    val = lead.get(key)
    if key == "weakness_flags" and isinstance(val, list):
        return ", ".join(val)
    return val if val is not None else ""


def save_leads_spreadsheet(leads: list[dict], output_dir: str = "./output") -> tuple[str, str]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    city_slug = (leads[0].get("city", "unknown") if leads else "unknown").lower().replace(" ", "_")
    type_slug = (leads[0].get("category", "businesses") if leads else "businesses").lower().replace(" ", "_")
    base = f"leads_{city_slug}_{type_slug}_{ts}"

    csv_path = out / f"{base}.csv"
    xlsx_path = out / f"{base}.xlsx"

    # --- CSV ---
    headers = [col[1] for col in COLUMNS]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for lead in leads:
            writer.writerow([_cell_value(lead, key) for key, _ in COLUMNS])

    # --- XLSX ---
    if not _OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor=_FILL_HEADER)
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    for row_idx, lead in enumerate(leads, start=2):
        score = lead.get("score", 0) or 0
        if score >= 7:
            row_color = _FILL_RED
        elif score >= 4:
            row_color = _FILL_YELLOW
        else:
            row_color = _FILL_GREEN
        fill = PatternFill("solid", fgColor=row_color)

        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(lead, key))
            cell.fill = fill

    # Auto-width
    col_widths = [len(h) for _, h in COLUMNS]
    for lead in leads:
        for i, (key, _) in enumerate(COLUMNS):
            val = str(_cell_value(lead, key))
            col_widths[i] = max(col_widths[i], min(len(val), 60))

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width + 2

    wb.save(str(xlsx_path))

    print(f"[export] Saved {len(leads)} leads → {xlsx_path} + {csv_path}")
    return str(xlsx_path), str(csv_path)
